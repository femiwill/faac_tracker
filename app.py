import os
import logging
from datetime import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
import requests as http_requests
from openpyxl import load_workbook
from io import BytesIO
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'faac-tracker-dev-key-change-in-prod')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///faac.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['ADMIN_PASSWORD'] = os.environ.get('ADMIN_PASSWORD', 'admin123')

db = SQLAlchemy(app)


# ── Models ──────────────────────────────────────────────────────────────────

class State(db.Model):
    __tablename__ = 'states'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    code = db.Column(db.String(10), nullable=False, unique=True)
    geo_zone = db.Column(db.String(50), nullable=False)
    lgas = db.relationship('LGA', backref='state', lazy=True)
    allocations = db.relationship('FAACAllocation', backref='state', lazy=True)
    igr_records = db.relationship('IGR', backref='state', lazy=True)


class LGA(db.Model):
    __tablename__ = 'lgas'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    state_id = db.Column(db.Integer, db.ForeignKey('states.id'), nullable=False)
    allocations = db.relationship('FAACAllocation', backref='lga', lazy=True)


class FAACAllocation(db.Model):
    __tablename__ = 'faac_allocations'
    id = db.Column(db.Integer, primary_key=True)
    state_id = db.Column(db.Integer, db.ForeignKey('states.id'), nullable=False)
    lga_id = db.Column(db.Integer, db.ForeignKey('lgas.id'), nullable=True)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    statutory_allocation = db.Column(db.Float, default=0)
    vat_allocation = db.Column(db.Float, default=0)
    total_gross = db.Column(db.Float, default=0)
    deductions = db.Column(db.Float, default=0)
    net_allocation = db.Column(db.Float, default=0)


class IGR(db.Model):
    __tablename__ = 'igr'
    id = db.Column(db.Integer, primary_key=True)
    state_id = db.Column(db.Integer, db.ForeignKey('states.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    quarter = db.Column(db.Integer, nullable=False)
    amount = db.Column(db.Float, default=0)


class ScrapeLog(db.Model):
    __tablename__ = 'scrape_logs'
    id = db.Column(db.Integer, primary_key=True)
    run_date = db.Column(db.DateTime, nullable=False)
    target_month = db.Column(db.Integer, nullable=False)
    target_year = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(20), nullable=False)  # 'success', 'failed', 'no_data'
    source = db.Column(db.String(100))  # 'nbs_excel', 'oagf', 'manual'
    states_added = db.Column(db.Integer, default=0)
    message = db.Column(db.Text)


# ── Helpers ─────────────────────────────────────────────────────────────────

MONTH_NAMES = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December'
}

GEO_ZONES = [
    'North Central', 'North East', 'North West',
    'South East', 'South South', 'South West'
]


def fmt_naira(amount):
    """Format amount in billions/millions for display."""
    if amount is None:
        return '₦0'
    if amount >= 1_000_000_000:
        return f'₦{amount / 1_000_000_000:,.2f}B'
    if amount >= 1_000_000:
        return f'₦{amount / 1_000_000:,.2f}M'
    return f'₦{amount:,.0f}'


app.jinja_env.filters['naira'] = fmt_naira
app.jinja_env.globals['MONTH_NAMES'] = MONTH_NAMES
app.jinja_env.globals['GEO_ZONES'] = GEO_ZONES


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


# ── Scraper ─────────────────────────────────────────────────────────────────

logger = logging.getLogger(__name__)

NBS_URL_PATTERNS = [
    'https://nigerianstat.gov.ng/resource/Disbursement%20{month},%20{year}.xlsx',
    'https://nigerianstat.gov.ng/resource/FAAC%20Disbursement%20{month}%20{year}.xlsx',
    'https://nigerianstat.gov.ng/resource/{month}%20{year}%20Disbursement.xlsx',
]


def _build_state_lookup():
    """Build a lookup dict mapping lowercase state name variants to State objects."""
    states = State.query.all()
    lookup = {}
    for s in states:
        lookup[s.name.lower()] = s
        lookup[s.name.lower().replace(' ', '')] = s
        # Handle FCT variations
        if s.name == 'FCT':
            lookup['fct abuja'] = s
            lookup['fct, abuja'] = s
            lookup['federal capital territory'] = s
    return lookup


def _try_download_nbs_excel(month_name, year):
    """Try each NBS URL pattern; return workbook or None."""
    for pattern in NBS_URL_PATTERNS:
        url = pattern.format(month=month_name, year=year)
        try:
            resp = http_requests.get(url, timeout=30)
            if resp.status_code == 200 and len(resp.content) > 1000:
                wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
                return wb, url
        except Exception:
            continue
    return None, None


def _parse_excel_data(wb, state_lookup):
    """Parse FAAC allocation data from an NBS Excel workbook.

    Returns list of dicts with keys: state_id, statutory, vat, deductions, net.
    """
    records = []
    ws = wb.active

    # Find the header row and column indices
    header_row = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        for cell in row:
            val = str(cell.value or '').strip().lower()
            if 'state' in val or 's/n' in val:
                header_row = cell.row
                break
        if header_row:
            # Map column headers
            for cell in row:
                val = str(cell.value or '').strip().lower()
                if 'statutory' in val:
                    col_map['statutory'] = cell.column - 1
                elif 'vat' in val:
                    col_map['vat'] = cell.column - 1
                elif 'deduction' in val:
                    col_map['deductions'] = cell.column - 1
                elif 'net' in val:
                    col_map['net'] = cell.column - 1
            break

    if not header_row:
        return records

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue

        cell_val = str(row[0]).strip().lower()
        # Skip summary/total rows
        if any(kw in cell_val for kw in ['total', 'grand', 'sum', 'note']):
            continue

        state = state_lookup.get(cell_val)
        if not state:
            # Try removing numbers/punctuation (e.g. "1. Abia" → "abia")
            cleaned = ''.join(c for c in cell_val if c.isalpha() or c == ' ').strip()
            state = state_lookup.get(cleaned)

        if not state:
            continue

        def safe_float(val):
            if val is None:
                return 0.0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        statutory = safe_float(row[col_map.get('statutory', 1)])
        vat = safe_float(row[col_map.get('vat', 2)])
        deductions = safe_float(row[col_map.get('deductions', 3)])
        net = safe_float(row[col_map.get('net', 4)])

        if net <= 0 and statutory <= 0:
            continue

        records.append({
            'state_id': state.id,
            'statutory': statutory,
            'vat': vat,
            'deductions': deductions,
            'net': net,
        })

    return records


def scrape_faac_data(target_month=None, target_year=None):
    """Scrape the latest FAAC allocation data from NBS.

    If target_month/year not specified, uses the previous month.
    """
    with app.app_context():
        now = datetime.utcnow()
        if target_month is None or target_year is None:
            # Target the previous month (data is released with a lag)
            if now.month == 1:
                target_month = 12
                target_year = now.year - 1
            else:
                target_month = now.month - 1
                target_year = now.year

        month_name = MONTH_NAMES[target_month]

        # Check if data already exists for this month
        existing = FAACAllocation.query.filter_by(
            month=target_month, year=target_year, lga_id=None
        ).first()
        if existing:
            log = ScrapeLog(
                run_date=now, target_month=target_month, target_year=target_year,
                status='no_data', source=None, states_added=0,
                message=f'Data for {month_name} {target_year} already exists in database.'
            )
            db.session.add(log)
            db.session.commit()
            logger.info(f'Scrape skipped: data for {month_name} {target_year} already exists.')
            return

        # Try NBS Excel download
        state_lookup = _build_state_lookup()
        wb, source_url = _try_download_nbs_excel(month_name, target_year)

        if wb:
            try:
                records = _parse_excel_data(wb, state_lookup)
                wb.close()

                if len(records) < 10:
                    log = ScrapeLog(
                        run_date=now, target_month=target_month, target_year=target_year,
                        status='failed', source='nbs_excel', states_added=0,
                        message=f'Excel found at {source_url} but only {len(records)} states parsed (expected 37). Data not inserted.'
                    )
                    db.session.add(log)
                    db.session.commit()
                    logger.warning(f'Scrape failed: only {len(records)} states parsed.')
                    return

                # Insert records
                for rec in records:
                    gross = rec['statutory'] + rec['vat']
                    alloc = FAACAllocation(
                        state_id=rec['state_id'], lga_id=None,
                        month=target_month, year=target_year,
                        statutory_allocation=rec['statutory'],
                        vat_allocation=rec['vat'],
                        total_gross=gross,
                        deductions=rec['deductions'],
                        net_allocation=rec['net'],
                    )
                    db.session.add(alloc)

                log = ScrapeLog(
                    run_date=now, target_month=target_month, target_year=target_year,
                    status='success', source='nbs_excel', states_added=len(records),
                    message=f'Successfully scraped {len(records)} state records from {source_url}.'
                )
                db.session.add(log)
                db.session.commit()
                logger.info(f'Scrape success: {len(records)} states for {month_name} {target_year}.')
                return

            except Exception as e:
                db.session.rollback()
                log = ScrapeLog(
                    run_date=now, target_month=target_month, target_year=target_year,
                    status='failed', source='nbs_excel', states_added=0,
                    message=f'Error parsing Excel from {source_url}: {str(e)}'
                )
                db.session.add(log)
                db.session.commit()
                logger.error(f'Scrape error: {e}')
                return

        # NBS not available
        log = ScrapeLog(
            run_date=now, target_month=target_month, target_year=target_year,
            status='no_data', source=None, states_added=0,
            message=f'No Excel file found for {month_name} {target_year} at NBS. Tried {len(NBS_URL_PATTERNS)} URL patterns.'
        )
        db.session.add(log)
        db.session.commit()
        logger.info(f'Scrape no_data: no file found for {month_name} {target_year}.')


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route('/terms')
def terms():
    return render_template('terms.html')


@app.route('/')
def index():
    states = State.query.order_by(State.name).all()

    # Latest month summary
    latest = db.session.query(
        FAACAllocation.year, FAACAllocation.month
    ).filter(
        FAACAllocation.lga_id.is_(None)
    ).order_by(
        FAACAllocation.year.desc(), FAACAllocation.month.desc()
    ).first()

    summary = []
    if latest:
        top_states = FAACAllocation.query.filter_by(
            year=latest.year, month=latest.month, lga_id=None
        ).order_by(FAACAllocation.net_allocation.desc()).limit(5).all()
        summary = top_states

    zones = {}
    for s in states:
        zones.setdefault(s.geo_zone, []).append(s)

    return render_template('index.html',
                           states=states, zones=zones,
                           summary=summary, latest=latest)


@app.route('/api/search')
def api_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])

    results = []
    states = State.query.filter(State.name.ilike(f'%{q}%')).limit(5).all()
    for s in states:
        results.append({'type': 'state', 'name': s.name, 'url': url_for('state_detail', name=s.name)})

    lgas = LGA.query.filter(LGA.name.ilike(f'%{q}%')).limit(5).all()
    for lg in lgas:
        results.append({'type': 'lga', 'name': f'{lg.name} ({lg.state.name})',
                        'url': url_for('lga_detail', state_name=lg.state.name, lga_name=lg.name)})

    return jsonify(results)


@app.route('/state/<name>')
def state_detail(name):
    state = State.query.filter(State.name.ilike(name)).first_or_404()

    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    alloc_query = FAACAllocation.query.filter_by(state_id=state.id, lga_id=None)
    if year:
        alloc_query = alloc_query.filter_by(year=year)
    if month:
        alloc_query = alloc_query.filter_by(month=month)

    allocations = alloc_query.order_by(
        FAACAllocation.year.desc(), FAACAllocation.month.desc()
    ).all()

    igr_data = IGR.query.filter_by(state_id=state.id).order_by(IGR.year.desc(), IGR.quarter).all()

    lga_allocations = []
    latest = db.session.query(
        FAACAllocation.year, FAACAllocation.month
    ).filter(
        FAACAllocation.state_id == state.id, FAACAllocation.lga_id.isnot(None)
    ).order_by(
        FAACAllocation.year.desc(), FAACAllocation.month.desc()
    ).first()

    if latest:
        lga_allocations = FAACAllocation.query.filter(
            FAACAllocation.state_id == state.id,
            FAACAllocation.lga_id.isnot(None),
            FAACAllocation.year == latest.year,
            FAACAllocation.month == latest.month
        ).order_by(FAACAllocation.net_allocation.desc()).all()

    available_years = db.session.query(FAACAllocation.year).filter_by(
        state_id=state.id, lga_id=None
    ).distinct().order_by(FAACAllocation.year.desc()).all()
    available_years = [y[0] for y in available_years]

    # Chart data
    chart_labels = [f"{MONTH_NAMES[a.month][:3]} {a.year}" for a in reversed(allocations)]
    chart_statutory = [a.statutory_allocation for a in reversed(allocations)]
    chart_vat = [a.vat_allocation for a in reversed(allocations)]
    chart_net = [a.net_allocation for a in reversed(allocations)]

    return render_template('state.html',
                           state=state, allocations=allocations,
                           igr_data=igr_data, lga_allocations=lga_allocations,
                           latest_lga=latest,
                           available_years=available_years,
                           filter_year=year, filter_month=month,
                           chart_labels=chart_labels,
                           chart_statutory=chart_statutory,
                           chart_vat=chart_vat,
                           chart_net=chart_net)


@app.route('/lga/<state_name>/<lga_name>')
def lga_detail(state_name, lga_name):
    state = State.query.filter(State.name.ilike(state_name)).first_or_404()
    lga = LGA.query.filter(LGA.name.ilike(lga_name), LGA.state_id == state.id).first_or_404()

    allocations = FAACAllocation.query.filter_by(lga_id=lga.id).order_by(
        FAACAllocation.year.desc(), FAACAllocation.month.desc()
    ).all()

    chart_labels = [f"{MONTH_NAMES[a.month][:3]} {a.year}" for a in reversed(allocations)]
    chart_net = [a.net_allocation for a in reversed(allocations)]

    return render_template('lga.html', state=state, lga=lga,
                           allocations=allocations,
                           chart_labels=chart_labels,
                           chart_net=chart_net)


@app.route('/compare', methods=['GET'])
def compare():
    states = State.query.order_by(State.name).all()
    selected_names = request.args.getlist('states')

    compared = []
    for name in selected_names[:3]:
        s = State.query.filter(State.name.ilike(name)).first()
        if s:
            allocs = FAACAllocation.query.filter_by(
                state_id=s.id, lga_id=None
            ).order_by(FAACAllocation.year, FAACAllocation.month).all()

            igr_total = db.session.query(db.func.sum(IGR.amount)).filter_by(
                state_id=s.id
            ).scalar() or 0

            compared.append({
                'state': s,
                'allocations': allocs,
                'igr_total': igr_total,
                'labels': [f"{MONTH_NAMES[a.month][:3]} {a.year}" for a in allocs],
                'net_values': [a.net_allocation for a in allocs],
            })

    return render_template('compare.html', states=states, compared=compared,
                           selected_names=selected_names)


# ── Admin ───────────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == app.config['ADMIN_PASSWORD']:
            session['admin'] = True
            flash('Logged in successfully.', 'success')
            return redirect(url_for('admin_dashboard'))
        flash('Incorrect password.', 'danger')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    flash('Logged out.', 'info')
    return redirect(url_for('index'))


@app.route('/admin')
@login_required
def admin_dashboard():
    states = State.query.order_by(State.name).all()
    scrape_logs = ScrapeLog.query.order_by(ScrapeLog.run_date.desc()).limit(20).all()
    try:
        next_run = scheduler.get_job('faac_monthly_scrape')
        next_run_time = next_run.next_run_time.strftime('%d %b %Y, %H:%M UTC') if next_run and next_run.next_run_time else None
    except Exception:
        next_run_time = None
    return render_template('admin.html', states=states, scrape_logs=scrape_logs,
                           next_run_time=next_run_time)


@app.route('/admin/add_allocation', methods=['POST'])
@login_required
def admin_add_allocation():
    state_id = request.form.get('state_id', type=int)
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    statutory = request.form.get('statutory_allocation', type=float, default=0)
    vat = request.form.get('vat_allocation', type=float, default=0)
    deductions = request.form.get('deductions', type=float, default=0)
    total_gross = statutory + vat
    net = total_gross - deductions

    existing = FAACAllocation.query.filter_by(
        state_id=state_id, lga_id=None, month=month, year=year
    ).first()

    if existing:
        existing.statutory_allocation = statutory
        existing.vat_allocation = vat
        existing.total_gross = total_gross
        existing.deductions = deductions
        existing.net_allocation = net
        flash('Allocation updated.', 'success')
    else:
        alloc = FAACAllocation(
            state_id=state_id, lga_id=None, month=month, year=year,
            statutory_allocation=statutory, vat_allocation=vat,
            total_gross=total_gross, deductions=deductions, net_allocation=net
        )
        db.session.add(alloc)
        flash('Allocation added.', 'success')

    db.session.commit()
    return redirect(url_for('admin_dashboard'))


@app.route('/api/lgas/<int:state_id>')
def api_lgas(state_id):
    lgas = LGA.query.filter_by(state_id=state_id).order_by(LGA.name).all()
    return jsonify([{'id': lg.id, 'name': lg.name} for lg in lgas])


@app.route('/admin/run_scraper', methods=['POST'])
@login_required
def admin_run_scraper():
    target_month = request.form.get('target_month', type=int)
    target_year = request.form.get('target_year', type=int)
    scrape_faac_data(target_month=target_month, target_year=target_year)
    flash('Scraper run completed. Check the scrape history below for results.', 'info')
    return redirect(url_for('admin_dashboard'))


# ── Init ────────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    # Auto-seed if database is empty (needed for Railway's ephemeral filesystem)
    if State.query.count() == 0:
        from seed_data import seed
        seed()

# ── Scheduler ───────────────────────────────────────────────────────────────

scheduler = BackgroundScheduler()
scheduler.add_job(
    func=scrape_faac_data,
    trigger='cron',
    day=15,
    hour=9,  # 9 AM UTC
    id='faac_monthly_scrape',
    misfire_grace_time=86400,  # allow 24h grace if missed
    replace_existing=True,
)
scheduler.start()
logger.info('APScheduler started. FAAC scraper scheduled for the 15th of each month at 9 AM UTC.')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
